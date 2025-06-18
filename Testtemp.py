import sys
import time
import requests
import json
import smtplib
import subprocess
import platform
import psutil
import socket
import os
import tempfile
import random
from datetime import datetime
from email.mime.text import MIMEText
from collections import deque
import xml.etree.ElementTree as ET
from xml.dom import minidom
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import io
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors

# --- PyQt5 Imports ---
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QVBoxLayout, QHBoxLayout, QPushButton,
    QWidget, QLabel, QFrame, QMessageBox, QComboBox, QLineEdit,
    QFormLayout, QGroupBox, QCheckBox, QProgressDialog, QTableWidget,
    QTableWidgetItem, QTabWidget, QSpinBox, QTextEdit, QHeaderView
)
from PyQt5.QtCore import QTimer, Qt, QThread, pyqtSignal, pyqtSlot
from PyQt5.QtGui import QFont, QColor

# --- Matplotlib Imports for plotting ---
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import matplotlib.dates as mdates

# --- Platform-specific imports ---
if platform.system() == 'Windows':
    try:
        import wmi
        import win32api
    except ImportError:
        print("Warning: WMI or py-win32 library not found. Some hardware info might be unavailable on Windows.")
        print("Try running: pip install wmi pywin32")
        wmi = None
else:
    wmi = None


class DataCollectorThread(QThread):
    data_updated = pyqtSignal(dict)
    monitoring_error = pyqtSignal(str, str)

    def __init__(self, poll_interval_ms, parent=None):
        super().__init__(parent)
        self.poll_interval_s = poll_interval_ms / 1000.0
        self._running = True
        self.wmi_instance = None
        self.error_reported = set()

    def stop(self):
        self._running = False

    def get_cpu_temperature(self, wmi_sensors):
        # Приоритет: WMI (OpenHardwareMonitor)
        if wmi_sensors:
            try:
                cpu_temps = {}
                for sensor in wmi_sensors:
                    if sensor.SensorType == 'Temperature' and 'cpu' in sensor.Name.lower():
                        cpu_temps[sensor.Name] = float(sensor.Value)
                if not cpu_temps: return None
                for name, value in cpu_temps.items():
                    if 'package' in name.lower(): return value
                return list(cpu_temps.values())[0]
            except Exception:
                return None

        # Резерв: psutil
        try:
            if hasattr(psutil, "sensors_temperatures"):
                temps = psutil.sensors_temperatures()
                if temps:
                    if 'coretemp' in temps: return temps['coretemp'][0].current
                    if 'k10temp' in temps: return temps['k10temp'][0].current
                    return list(temps.values())[0][0].current
        except Exception:
            pass
        return None

    def get_gpu_info(self, wmi_sensors):
        # Приоритет: WMI (OpenHardwareMonitor)
        if wmi_sensors:
            try:
                gpu_info = {}
                for sensor in wmi_sensors:
                    if 'gpu' in sensor.Name.lower():
                        if sensor.SensorType == 'Temperature': gpu_info['temp'] = float(sensor.Value)
                        if sensor.SensorType == 'Load' and 'core' in sensor.Name.lower(): gpu_info['load'] = float(
                            sensor.Value)
                if gpu_info: return gpu_info
            except Exception:
                pass

        # Резерв: nvidia-smi
        try:
            startupinfo = None
            if platform.system() == "Windows":
                startupinfo = subprocess.STARTUPINFO()
                startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
            output = subprocess.check_output(
                ['nvidia-smi', '--query-gpu=utilization.gpu,temperature.gpu,memory.used,memory.total',
                 '--format=csv,noheader,nounits'], stderr=subprocess.DEVNULL, text=True, startupinfo=startupinfo)
            v = output.strip().split(',')
            return {'load': float(v[0]), 'temp': float(v[1]), 'mem_used': float(v[2]), 'mem_total': float(v[3])}
        except Exception:
            pass
        return None

    def run(self):
        # Инициализируем WMI здесь, ВНУТРИ рабочего потока.
        if platform.system() == 'Windows':
            try:
                import wmi
                self.wmi_instance = wmi.WMI(namespace="root\\OpenHardwareMonitor")
                print("[INFO] WMI подключение успешно создано в рабочем потоке.")
            except Exception as e:
                print(f"[ERROR] Не удалось инициализировать WMI в потоке. Данные с OHM недоступны. Ошибка: {e}")
                self.wmi_instance = None

        while self._running:
            wmi_sensors_list = None
            if self.wmi_instance:
                try:
                    wmi_sensors_list = self.wmi_instance.Sensor()
                except Exception as e:
                    print(f"[ERROR] Ошибка при опросе сенсоров WMI: {e}")

            data_bundle = {'timestamp': datetime.now()}

            cpu_temp = self.get_cpu_temperature(wmi_sensors_list)
            gpu_data = self.get_gpu_info(wmi_sensors_list)


            try:
                data_bundle['cpu'] = {
                    'percent': psutil.cpu_percent(interval=None),
                    'frequency': psutil.cpu_freq()._asdict() if psutil.cpu_freq() else None,
                    'cores_physical': psutil.cpu_count(logical=False),
                    'cores_logical': psutil.cpu_count(logical=True),
                    'temperature': cpu_temp
                }
            except Exception as e:
                print(f"Error collecting CPU data: {e}")
            try:
                data_bundle['memory'] = {
                    'virtual': psutil.virtual_memory()._asdict(),
                    'swap': psutil.swap_memory()._asdict()
                }
            except Exception as e:
                if 'PdhAddEnglishCounterW' in str(e) and 'memory' not in self.error_reported:
                    self.monitoring_error.emit('Memory', str(e))
                    self.error_reported.add('memory')
                else:
                    pass
            try:
                partitions = psutil.disk_partitions(all=False)
                disk_data = {}
                for part in partitions:
                    if not (('fixed' in part.opts if platform.system() == 'Windows' else part.device.startswith(
                            ('/dev/sd', '/dev/nvme')))):
                        continue
                    try:
                        disk_data[part.mountpoint.replace(":", "_drive")] = psutil.disk_usage(part.mountpoint)._asdict()
                    except Exception:
                        continue
                data_bundle['disk'] = disk_data
            except Exception as e:
                print(f"Error collecting Disk data: {e}")

            data_bundle['gpu'] = gpu_data

            try:
                net_io_pernic = psutil.net_io_counters(pernic=True)
                data_bundle['network'] = {k.replace(":", "_").replace(" ", "_"): v._asdict() for k, v in
                                          net_io_pernic.items()}
            except Exception as e:
                print(f"Error collecting Network data: {e}")

            self.data_updated.emit(data_bundle)
            time.sleep(self.poll_interval_s)

class SpeedTestThread(QThread):
    result_ready = pyqtSignal(str)

    def run(self):
        try:
            self.result_ready.emit("Running speed test...")
            QApplication.processEvents()
            test_file_url = "http://speedtest.tele2.net/1MB.zip"
            start_time = time.time()
            with requests.get(test_file_url, stream=True, timeout=20) as r:
                r.raise_for_status()
                total_length = 0
                for chunk in r.iter_content(chunk_size=8192):
                    if chunk: total_length += len(chunk)
            duration = time.time() - start_time
            if duration > 0:
                speed_mbps = (total_length * 8) / (duration * 1024 * 1024)
                result_text = (f"Downloaded {total_length / 1024:.2f} KB in {duration:.2f} seconds\n"
                               f"Download Speed: {speed_mbps:.2f} Mbps")
            else:
                result_text = "Speed test failed: duration was zero."
            self.result_ready.emit(result_text)
        except requests.RequestException as e:
            self.result_ready.emit(f"Speed test failed: Network error\n{str(e)}")
        except Exception as e:
            self.result_ready.emit(f"Speed test failed: An unexpected error occurred.\n{str(e)}")


class SystemMonitorApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("System Monitor")
        self.setGeometry(100, 100, 1400, 900)
        self.monitoring_active = {'cpu': True, 'memory': True, 'disk': True, 'gpu': True, 'network': True}
        self.max_graph_points = 100
        self.time_points = deque(maxlen=self.max_graph_points)
        self.cpu_usage_points = deque(maxlen=self.max_graph_points)
        self.cpu_temp_points = deque(maxlen=self.max_graph_points)
        self.gpu_load_points = deque(maxlen=self.max_graph_points)
        self.gpu_temp_points = deque(maxlen=self.max_graph_points)
        self.mem_usage_points = deque(maxlen=self.max_graph_points)
        self.historical_data = deque(maxlen=2000)
        self.last_net_io = None
        self.last_update_time = None
        self.simulated_devices = []
        self.alert_history = deque(maxlen=100)
        self.last_alert_time = {}
        self.alerts_enabled = True
        self.initialized_tabs = set()
        self.speed_test_thread = SpeedTestThread()
        self.settings = {
            'poll_interval': 2000, 'cpu_temp_threshold': 80, 'gpu_temp_threshold': 85,
            'ram_threshold': 90, 'disk_threshold': 90,
            'email_notifications': False, 'email_server': 'smtp.example.com', 'email_port': 587,
            'email_from': 'monitor@example.com', 'email_to': 'admin@example.com', 'email_password': '',
            'popup_alerts': True
        }
        self.create_tabs()
        self.init_ui()
        self.load_settings()
        self.init_monitoring()
        self.aux_timer = QTimer(self)
        self.aux_timer.timeout.connect(self.update_simulated_devices_view)
        self.aux_timer.start(3000)

    def create_tabs(self):
        self.dashboard_tab, self.cpu_tab, self.memory_tab, self.disk_tab, self.gpu_tab = QWidget(), QWidget(), QWidget(), QWidget(), QWidget()
        self.network_tab, self.multi_device_tab, self.alerts_tab, self.reports_tab = QWidget(), QWidget(), QWidget(), QWidget()
        self.settings_tab, self.tools_tab = QWidget(), QWidget()

    def init_ui(self):
        self.main_widget = QWidget()
        self.setCentralWidget(self.main_widget)
        self.main_layout = QHBoxLayout(self.main_widget)
        self.main_layout.setContentsMargins(0, 0, 0, 0)
        self.main_layout.setSpacing(0)
        self.init_sidebar()
        self.init_content_area()
        self.statusBar().showMessage("Используется")

    def init_sidebar(self):
        self.sidebar = QFrame()
        self.sidebar.setFixedWidth(200)
        self.sidebar.setStyleSheet("background-color: #2c3e50; color: white;")
        self.sidebar_layout = QVBoxLayout(self.sidebar)
        self.sidebar_layout.setContentsMargins(10, 10, 10, 10)
        self.sidebar_layout.setSpacing(8)
        buttons = [
            ("Главный экран", lambda: self.tabs.setCurrentIndex(0)), ("ЦПУ", lambda: self.tabs.setCurrentIndex(1)),
            ("ОЗУ", lambda: self.tabs.setCurrentIndex(2)), ("Диски", lambda: self.tabs.setCurrentIndex(3)),
            ("ГПУ", lambda: self.tabs.setCurrentIndex(4)), ("Сеть", lambda: self.tabs.setCurrentIndex(5)),
            ("Несколько устройств", lambda: self.tabs.setCurrentIndex(6)), ("Уведомления", lambda: self.tabs.setCurrentIndex(7)),
            ("Отчеты", lambda: self.tabs.setCurrentIndex(8)), ("Настройки", lambda: self.tabs.setCurrentIndex(9)),
            ("Инструменты", lambda: self.tabs.setCurrentIndex(10))
        ]
        for text, handler in buttons:
            btn = QPushButton(text)
            btn.setStyleSheet("""
                QPushButton { background-color: #34495e; color: white; border: none; padding: 12px; text-align: left; border-radius: 5px; }
                QPushButton:hover { background-color: #4a6b8a; }
                QPushButton:pressed { background-color: #5a82a8; }
            """)
            btn.clicked.connect(handler)
            self.sidebar_layout.addWidget(btn)
        self.sidebar_layout.addStretch()
        self.main_layout.addWidget(self.sidebar)

    def init_content_area(self):
        self.content_area = QFrame()
        self.content_area.setStyleSheet("background-color: #ecf0f1;")
        self.content_layout = QVBoxLayout(self.content_area)
        self.tabs = QTabWidget()
        self.tabs.tabBar().setVisible(False)
        tab_map = {
            "Главный экран": self.dashboard_tab, "ЦПУ": self.cpu_tab, "ОЗУ": self.memory_tab, "Диски": self.disk_tab,
            "ГПУ": self.gpu_tab, "Сеть": self.network_tab, "Несколько устройств": self.multi_device_tab,
            "Уведомления": self.alerts_tab, "Отчеты": self.reports_tab, "Настройки": self.settings_tab,
            "Инструменты": self.tools_tab
        }
        for name, widget in tab_map.items(): self.tabs.addTab(widget, name)
        self.tabs.currentChanged.connect(self.on_tab_changed)
        self.content_layout.addWidget(self.tabs)
        self.main_layout.addWidget(self.content_area)
        self.on_tab_changed(0)

    def on_tab_changed(self, index):
        if index in self.initialized_tabs:
            self.update_views_for_tab(index)
            return
        init_functions = {
            0: self.init_dashboard, 1: self.init_cpu_tab, 2: self.init_memory_tab, 3: self.init_disk_tab,
            4: self.init_gpu_tab, 5: self.init_network_tab, 6: self.init_multi_device_tab,
            7: self.init_alerts_tab, 8: self.init_reports_tab, 9: self.init_settings_tab, 10: self.init_tools_tab
        }
        if index in init_functions:
            init_functions[index]()
            self.initialized_tabs.add(index)
            self.update_views_for_tab(index)

    def update_views_for_tab(self, index):
        if not self.historical_data: return
        update_map = {
            0: lambda: (self.update_cpu_chart(), self.update_gpu_dashboard_chart(),
                        self.update_dashboard_alerts_view()),
            1: self.update_cpu_view, 2: self.update_memory_view, 3: self.update_disk_view,
            4: self.update_gpu_view, 5: self.update_network_view, 7: self.update_alerts_history_view
        }
        if index in update_map: update_map[index]()

    def init_monitoring(self):
        poll_interval = self.settings.get('poll_interval', 2000)
        self.data_collector = DataCollectorThread(poll_interval, self)
        self.data_collector.data_updated.connect(self.handle_data_update)
        self.data_collector.monitoring_error.connect(self.handle_monitoring_error)
        self.data_collector.start()

    @pyqtSlot(str, str)
    def handle_monitoring_error(self, component, error_message):
        component_key = component.lower()
        if not self.monitoring_active.get(component_key, True): return
        self.monitoring_active[component_key] = False
        QMessageBox.critical(self, f"Monitoring Error: {component}",
                             f"Failed to collect {component} data. This component will be disabled.\n\n"
                             f"Error: {error_message}\n\n"
                             "On Windows, this can often be fixed by running the following command in an "
                             "administrator command prompt and then restarting the application:\n\n"
                             "lodctr /r")

    @pyqtSlot(dict)
    def handle_data_update(self, data):
        self.historical_data.append(data)
        now = data['timestamp']
        self.time_points.append(now)
        cpu_data = data.get('cpu', {})
        self.cpu_usage_points.append(cpu_data.get('percent'))
        self.cpu_temp_points.append(cpu_data.get('temperature'))
        mem_data = data.get('memory', {}).get('virtual', {})
        self.mem_usage_points.append(mem_data.get('percent'))
        gpu_info = data.get('gpu') or {}
        self.gpu_load_points.append(gpu_info.get('load'))
        self.gpu_temp_points.append(gpu_info.get('temp'))
        if 0 in self.initialized_tabs:
            cpu_percent = cpu_data.get('percent', 0)
            cpu_temp = cpu_data.get('temperature')
            temp_str = f"{cpu_temp:.0f}°C" if cpu_temp is not None else "N/A"
            self.cpu_overview.setText(f"<b>CPU:</b> {cpu_percent:.1f}% | {temp_str}")
            if self.monitoring_active['memory'] and mem_data:
                self.memory_overview.setText(
                    f"<b>Memory:</b> {mem_data.get('percent', 0)}% ({mem_data.get('used', 0) / 1e9:.1f}/{mem_data.get('total', 0) / 1e9:.1f} GB)")
            else:
                self.memory_overview.setText(
                    "<b>Memory:</b> Disabled" if not self.monitoring_active['memory'] else "<b>Memory:</b> Error")
            disk_usage_map, root_disk = data.get('disk', {}), os.path.abspath(os.sep)
            disk_usage = disk_usage_map.get(root_disk.replace(":", "_drive"))
            self.disk_overview.setText(
                f"<b>Disk ({root_disk}):</b> {disk_usage['percent']}%" if disk_usage else "<b>Disk:</b> N/A")
            load_str = f"{gpu_info.get('load', 0):.0f}%" if gpu_info and 'load' in gpu_info and gpu_info[
                'load'] is not None else "N/A"
            temp_str = f"{gpu_info.get('temp', 0):.0f}°C" if gpu_info and 'temp' in gpu_info and gpu_info[
                'temp'] is not None else "N/A"
            self.gpu_overview.setText(f"<b>GPU:</b> {load_str} | {temp_str}")
        self.check_for_alerts(data)
        self.update_views_for_tab(self.tabs.currentIndex())

    def check_for_alerts(self, data):
        cpu_temp = data.get('cpu', {}).get('temperature')
        if cpu_temp and cpu_temp > self.settings['cpu_temp_threshold']:
            self.trigger_alert('CPU', f"High temperature: {cpu_temp:.0f}°C")
        mem_percent = data.get('memory', {}).get('virtual', {}).get('percent')
        if self.monitoring_active['memory'] and mem_percent and mem_percent > self.settings['ram_threshold']:
            self.trigger_alert('Memory', f'High usage: {mem_percent:.0f}%')
        gpu_info = data.get('gpu')
        if gpu_info and gpu_info.get('temp') and gpu_info.get('temp') > self.settings['gpu_temp_threshold']:
            self.trigger_alert('GPU', f"High temperature: {gpu_info.get('temp'):.0f}°C")
        for mount, usage in data.get('disk', {}).items():
            if usage['percent'] > self.settings['disk_threshold']:
                self.trigger_alert('Disk', f"High usage on {mount.replace('_drive', ':')}: {usage['percent']:.0f}%")

    def init_dashboard(self):
        layout = QVBoxLayout(self.dashboard_tab)
        overview_group, overview_layout = QGroupBox("Системные показатели"), QHBoxLayout()
        self.cpu_overview, self.memory_overview, self.disk_overview, self.gpu_overview = [
            QLabel(f"<b>{t}:</b> Loading...") for t in ["CPU", "Memory", "Disk", "GPU"]]
        for label in [self.cpu_overview, self.memory_overview, self.disk_overview,
                      self.gpu_overview]: overview_layout.addWidget(label)
        overview_group.setLayout(overview_layout)
        layout.addWidget(overview_group)
        charts_layout = QHBoxLayout()
        self.cpu_fig, self.cpu_ax = Figure(figsize=(5, 3), dpi=100), None
        self.cpu_canvas = FigureCanvas(self.cpu_fig)
        self.gpu_dashboard_fig, self.gpu_dashboard_ax = Figure(figsize=(5, 3), dpi=100), None
        self.gpu_dashboard_canvas = FigureCanvas(self.gpu_dashboard_fig)
        charts_layout.addWidget(self.cpu_canvas);
        charts_layout.addWidget(self.gpu_dashboard_canvas)
        layout.addLayout(charts_layout)
        self.cpu_ax = self.cpu_fig.add_subplot(111)
        self.cpu_usage_line, = self.cpu_ax.plot([], [], color='tab:blue', label='Usage')
        self.cpu_ax2 = self.cpu_ax.twinx()
        self.cpu_temp_line, = self.cpu_ax2.plot([], [], color='tab:red', label='ТЕМП')
        self.setup_chart_axes(self.cpu_ax, self.cpu_ax2, 'Используется (%)', 'ТЕМП (°C)', "Мониторинг ЦПУ")
        self.gpu_dashboard_ax = self.gpu_dashboard_fig.add_subplot(111)
        self.gpu_load_line, = self.gpu_dashboard_ax.plot([], [], color='tab:green', label='Нагрузка')
        self.gpu_ax2_dashboard = self.gpu_dashboard_ax.twinx()
        self.gpu_temp_line_dashboard, = self.gpu_ax2_dashboard.plot([], [], color='tab:orange', label='ТЕМП')
        self.setup_chart_axes(self.gpu_dashboard_ax, self.gpu_ax2_dashboard, 'Используется (%)', 'ТЕМП (°C)', "ГПУ Мониторинг")
        alerts_group, alerts_layout = QGroupBox("недавние оповещения"), QVBoxLayout()
        self.alerts_table_dashboard = QTableWidget(5, 3)
        self.alerts_table_dashboard.setHorizontalHeaderLabels(["Время", "Компонент", "Сообщение"])
        self.alerts_table_dashboard.horizontalHeader().setStretchLastSection(True)
        self.alerts_table_dashboard.setEditTriggers(QTableWidget.NoEditTriggers)
        alerts_layout.addWidget(self.alerts_table_dashboard)
        alerts_group.setLayout(alerts_layout)
        layout.addWidget(alerts_group)

    def setup_chart_axes(self, ax1, ax2, label1, label2, title):
        ax1.set_title(title);
        ax1.set_ylabel(label1, color=ax1.get_lines()[0].get_color())
        ax1.tick_params(axis='y', labelcolor=ax1.get_lines()[0].get_color())
        ax1.set_ylim(0, 105);
        ax1.grid(True, linestyle='--', alpha=0.6)
        ax2.set_ylabel(label2, color=ax2.get_lines()[0].get_color())
        ax2.tick_params(axis='y', labelcolor=ax2.get_lines()[0].get_color())
        ax2.set_ylim(20, 105)
        ax1.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M:%S'))
        ax1.figure.tight_layout()

    def _update_single_chart(self, line1, x1_data, y1_data, line2, x2_data, y2_data, ax, canvas):
        valid_points1 = [(t, v) for t, v in zip(x1_data, y1_data) if v is not None]
        if valid_points1:
            line1.set_data(*zip(*valid_points1))
        else:
            line1.set_data([], [])
        valid_points2 = [(t, v) for t, v in zip(x2_data, y2_data) if v is not None]
        if valid_points2:
            line2.set_data(*zip(*valid_points2))
        else:
            line2.set_data([], [])
        # Устанавливаем границы оси X, только если у нас есть диапазон (больше 1 точки)
        if len(x1_data) > 1:
            ax.set_xlim(x1_data[0], x1_data[-1])

        ax.figure.canvas.draw_idle()

    def update_cpu_chart(self):
        self._update_single_chart(self.cpu_usage_line, self.time_points, self.cpu_usage_points,
                                  self.cpu_temp_line, self.time_points, self.cpu_temp_points,
                                  self.cpu_ax, self.cpu_canvas)

    def update_gpu_dashboard_chart(self):
        self._update_single_chart(self.gpu_load_line, self.time_points, self.gpu_load_points,
                                  self.gpu_temp_line_dashboard, self.time_points, self.gpu_temp_points,
                                  self.gpu_dashboard_ax, self.gpu_dashboard_canvas)

    def init_cpu_tab(self):
        layout = QVBoxLayout(self.cpu_tab)
        self.cpu_detail_fig = Figure(figsize=(10, 5), dpi=100)
        self.cpu_detail_canvas = FigureCanvas(self.cpu_detail_fig)
        layout.addWidget(self.cpu_detail_canvas)
        self.cpu_detail_ax = self.cpu_detail_fig.add_subplot(111)
        self.cpu_detail_usage_line, = self.cpu_detail_ax.plot([], [], color='tab:blue', label='Usage')
        self.cpu_detail_ax2 = self.cpu_detail_ax.twinx()
        self.cpu_detail_temp_line, = self.cpu_detail_ax2.plot([], [], color='tab:red', label='Temp')
        self.setup_chart_axes(self.cpu_detail_ax, self.cpu_detail_ax2, 'ЦПУ (%)', 'ТЕМП (°C)', "ЦПУ ")
        self.cpu_table = QTableWidget(1, 4)
        self.cpu_table.setHorizontalHeaderLabels(["Ядра (Физические/Потоковые)", "Текущая скорость", "Макс скорость", "Использовано"])
        self.cpu_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(self.cpu_table)

    def update_cpu_view(self):
        if not self.historical_data: return
        data = self.historical_data[-1].get('cpu', {})
        freq = data.get('frequency', {})
        self._update_single_chart(self.cpu_detail_usage_line, self.time_points, self.cpu_usage_points,
                                  self.cpu_detail_temp_line, self.time_points, self.cpu_temp_points,
                                  self.cpu_detail_ax, self.cpu_detail_canvas)
        self.cpu_table.setItem(0, 0, QTableWidgetItem(
            f"{data.get('cores_physical', 'N/A')}/{data.get('cores_logical', 'N/A')}"))
        self.cpu_table.setItem(0, 1, QTableWidgetItem(f"{freq.get('current', 0):.2f} MHz" if freq else "N/A"))
        self.cpu_table.setItem(0, 2, QTableWidgetItem(f"{freq.get('max', 0):.2f} MHz" if freq else "N/A"))
        self.cpu_table.setItem(0, 3, QTableWidgetItem(f"{data.get('percent', 0):.1f}%"))

    def init_memory_tab(self):
        layout = QVBoxLayout(self.memory_tab)
        self.memory_detail_fig, self.memory_detail_ax = Figure(figsize=(10, 5), dpi=100), None
        self.memory_detail_canvas = FigureCanvas(self.memory_detail_fig)
        self.memory_detail_ax = self.memory_detail_fig.add_subplot(111)
        layout.addWidget(self.memory_detail_canvas)
        self.mem_usage_line, = self.memory_detail_ax.plot([], [], 'g-', label="ОЗУ Используется(%)")
        self.memory_detail_ax.set_ylim(0, 105);
        self.memory_detail_ax.set_ylabel("Usage (%)")
        self.memory_detail_ax.grid(True);
        self.memory_detail_ax.legend();
        self.memory_detail_fig.tight_layout()
        self.memory_table = QTableWidget(2, 4)
        self.memory_table.setHorizontalHeaderLabels(["Всего", "Используется", "Свободно", "Исп в %"])
        self.memory_table.setVerticalHeaderLabels(["RAM", "Swap"])
        self.memory_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(self.memory_table)

    def update_memory_view(self):
        if not self.historical_data: return
        data = self.historical_data[-1].get('memory', {})
        mem, swap = data.get('virtual'), data.get('swap')
        if not self.monitoring_active['memory']:
            self.memory_table.setRowCount(1);
            self.memory_table.setColumnCount(1)
            self.memory_table.setSpan(0, 0, 2, 4)
            self.memory_table.setItem(0, 0, QTableWidgetItem("Monitoring disabled due to system error."))
            return
        self.memory_table.setRowCount(2);
        self.memory_table.setColumnCount(4);
        self.memory_table.setSpan(0, 0, 1, 1)
        for i, m in enumerate([mem, swap]):
            if m:
                self.memory_table.setItem(i, 0, QTableWidgetItem(f"{m.get('total', 0) / 1e9:.2f} GB"))
                self.memory_table.setItem(i, 1, QTableWidgetItem(f"{m.get('used', 0) / 1e9:.2f} GB"))
                self.memory_table.setItem(i, 2, QTableWidgetItem(f"{m.get('free', 0) / 1e9:.2f} GB"))
                self.memory_table.setItem(i, 3, QTableWidgetItem(f"{m.get('percent', 0)}%"))
        valid_points = [(t, v) for t, v in zip(self.time_points, self.mem_usage_points) if v is not None]
        if valid_points:
            self.mem_usage_line.set_data(*zip(*valid_points))
        else:
            self.mem_usage_line.set_data([], [])
        if self.time_points: self.memory_detail_ax.set_xlim(self.time_points[0], self.time_points[-1])
        self.memory_detail_canvas.draw_idle()

    def init_disk_tab(self):
        layout = QVBoxLayout(self.disk_tab)
        self.disk_table = QTableWidget(0, 5)
        self.disk_table.setHorizontalHeaderLabels(["Название", "Всего", "Занято", "Свободно", "Используется в %"])
        self.disk_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.disk_table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.disk_table)

    def update_disk_view(self):
        if not self.historical_data: return
        disk_usages = self.historical_data[-1].get('disk', {})
        self.disk_table.setRowCount(len(disk_usages))
        for i, (mount, usage) in enumerate(disk_usages.items()):
            self.disk_table.setItem(i, 0, QTableWidgetItem(mount.replace('_drive', ':')))
            self.disk_table.setItem(i, 1, QTableWidgetItem(f"{usage.get('total', 0) / 1e9:.2f} GB"))
            self.disk_table.setItem(i, 2, QTableWidgetItem(f"{usage.get('used', 0) / 1e9:.2f} GB"))
            self.disk_table.setItem(i, 3, QTableWidgetItem(f"{usage.get('free', 0) / 1e9:.2f} GB"))
            self.disk_table.setItem(i, 4, QTableWidgetItem(f"{usage.get('percent', 0):.1f}%"))

    def init_gpu_tab(self):
        layout = QVBoxLayout(self.gpu_tab)
        self.gpu_detail_fig = Figure(figsize=(10, 5), dpi=100)
        self.gpu_detail_canvas = FigureCanvas(self.gpu_detail_fig)
        layout.addWidget(self.gpu_detail_canvas)
        self.gpu_detail_ax = self.gpu_detail_fig.add_subplot(111)
        self.gpu_detail_load_line, = self.gpu_detail_ax.plot([], [], color='tab:green', label='Используется')
        self.gpu_detail_ax2 = self.gpu_detail_ax.twinx()
        self.gpu_detail_temp_line, = self.gpu_detail_ax2.plot([], [], color='tab:orange', label='Темп')
        self.setup_chart_axes(self.gpu_detail_ax, self.gpu_detail_ax2, 'ГПУ (%)', 'Темп (°C)', "ГПУ")
        self.gpu_table = QTableWidget(0, 2)
        self.gpu_table.setHorizontalHeaderLabels(["Метрика", "Показатель"])
        self.gpu_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(self.gpu_table)

    def update_gpu_view(self):
        if not self.historical_data: return
        gpu_info = self.historical_data[-1].get('gpu')
        self._update_single_chart(self.gpu_detail_load_line, self.time_points, self.gpu_load_points,
                                  self.gpu_detail_temp_line, self.time_points, self.gpu_temp_points,
                                  self.gpu_detail_ax, self.gpu_detail_canvas)
        if gpu_info:
            self.gpu_table.setRowCount(len(gpu_info))
            for i, (k, v) in enumerate(gpu_info.items()):
                self.gpu_table.setItem(i, 0, QTableWidgetItem(k.replace('_', ' ').capitalize()))
                self.gpu_table.setItem(i, 1, QTableWidgetItem(str(v)))
        else:
            self.gpu_table.setRowCount(1)
            self.gpu_table.setItem(0, 0, QTableWidgetItem("ГПУ информация"));
            self.gpu_table.setItem(0, 1, QTableWidgetItem("Не определенно"))

    def init_network_tab(self):
        layout = QVBoxLayout(self.network_tab)
        self.network_table = QTableWidget(0, 5)
        self.network_table.setHorizontalHeaderLabels(
            ["Интерфейс", "Отправлено(Всего)", "Принято(Всего)", "Скорость отправки", "Скорость скачивания"])
        self.network_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.network_table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.network_table)

    def update_network_view(self):
        if not self.historical_data: return
        current_data = self.historical_data[-1]
        current_time, current_io = current_data['timestamp'], current_data.get('network')
        if not current_io: return
        self.network_table.setRowCount(len(current_io))
        for i, (iface, counters) in enumerate(current_io.items()):
            self.network_table.setItem(i, 0, QTableWidgetItem(iface.replace('_', ' ')))
            self.network_table.setItem(i, 1, QTableWidgetItem(f"{counters.get('bytes_sent', 0) / 1e9:.3f} GB"))
            self.network_table.setItem(i, 2, QTableWidgetItem(f"{counters.get('bytes_recv', 0) / 1e9:.3f} GB"))
            sent_rate_str, recv_rate_str = "N/A", "N/A"
            if self.last_net_io and self.last_update_time:
                time_delta = (current_time - self.last_update_time).total_seconds()
                if time_delta > 0 and iface in self.last_net_io:
                    last_counters = self.last_net_io[iface]
                    sent_rate = (counters.get('bytes_sent', 0) - last_counters.get('bytes_sent', 0)) / time_delta
                    recv_rate = (counters.get('bytes_recv', 0) - last_counters.get('bytes_recv', 0)) / time_delta
                    sent_rate_str = f"{sent_rate / 1024:.2f} KB/s"
                    recv_rate_str = f"{recv_rate / 1024:.2f} KB/s"
            self.network_table.setItem(i, 3, QTableWidgetItem(sent_rate_str))
            self.network_table.setItem(i, 4, QTableWidgetItem(recv_rate_str))
        self.last_net_io, self.last_update_time = current_io, current_time

    def discover_network_devices(self):
        self.statusBar().showMessage(" Обноружение устройства в сети");
        self.populate_simulated_devices()
        QTimer.singleShot(2000, lambda: self.statusBar().showMessage("Готово"))

    def init_multi_device_tab(self):
        layout, button_layout = QVBoxLayout(self.multi_device_tab), QHBoxLayout()
        discover_btn = QPushButton(" Обновить");
        discover_btn.clicked.connect(self.discover_network_devices)
        button_layout.addWidget(discover_btn);
        button_layout.addStretch();
        layout.addLayout(button_layout)
        self.multi_device_table = QTableWidget(0, 5)
        self.multi_device_table.setHorizontalHeaderLabels(["Имя устройства", "IP ", "Статус", "ЦПУ %", "ОЗУ %"])
        self.multi_device_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.multi_device_table.setEditTriggers(QTableWidget.NoEditTriggers);
        self.multi_device_table.setSortingEnabled(True)
        layout.addWidget(self.multi_device_table);
        self.discover_network_devices()

    def populate_simulated_devices(self):
        self.simulated_devices.clear()
        for i, name in enumerate(["DesktopCova", "LaptopCova", "LaptopLika", "NONAME", "SERVER"]):
            self.simulated_devices.append(
                {'name': name, 'ip': f'192.168.1.{10 + i}', 'status': 'Online', 'cpu': random.randint(5, 70),
                 'ram': random.randint(20, 80)})
        self.update_simulated_devices_view()

    def update_simulated_devices_view(self):
        if 6 not in self.initialized_tabs or not self.simulated_devices: return
        for device in self.simulated_devices:
            if device['status'] == 'Online':
                device['cpu'] = max(0, min(100, device['cpu'] + random.randint(-5, 5)))
                device['ram'] = max(0, min(100, device['ram'] + random.randint(-3, 3)))
                if random.random() < 0.01: device['status'] = 'Offline'
            elif random.random() < 0.05:
                device['status'] = 'Online'
        self.multi_device_table.setSortingEnabled(False)
        self.multi_device_table.setRowCount(len(self.simulated_devices))
        for i, device in enumerate(self.simulated_devices):
            self.multi_device_table.setItem(i, 0, QTableWidgetItem(device['name']))
            self.multi_device_table.setItem(i, 1, QTableWidgetItem(device['ip']))
            status_item = QTableWidgetItem(device['status'])
            status_item.setForeground(QColor('green') if device['status'] == 'Online' else QColor('red'))
            self.multi_device_table.setItem(i, 2, status_item)
            self.multi_device_table.setItem(i, 3, QTableWidgetItem(str(device['cpu'])))
            self.multi_device_table.setItem(i, 4, QTableWidgetItem(str(device['ram'])))
        self.multi_device_table.setSortingEnabled(True)

    def trigger_alert(self, component, message):
        if not self.alerts_enabled: return
        alert_key, now = f"{component}-{message.split(':')[0]}", time.time()
        if now - self.last_alert_time.get(alert_key, 0) < 300: return
        self.last_alert_time[alert_key] = now
        self.alert_history.appendleft(
            {'Время': datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 'Компонент': component, 'Сообщение': message})
        self.update_dashboard_alerts_view()
        if 7 in self.initialized_tabs: self.update_alerts_history_view()
        if self.settings.get('Всплывающие оповещения', True): QMessageBox.warning(self, f"{component} уведомление", message)

    def update_dashboard_alerts_view(self):
        if 0 in self.initialized_tabs:
            self.alerts_table_dashboard.clearContents()
            for i, alert in enumerate(list(self.alert_history)[:5]):
                self.alerts_table_dashboard.setItem(i, 0, QTableWidgetItem(alert['Время']))
                self.alerts_table_dashboard.setItem(i, 1, QTableWidgetItem(alert['Компонент']))
                self.alerts_table_dashboard.setItem(i, 2, QTableWidgetItem(alert['Сообщение']))

    def init_alerts_tab(self):
        layout = QVBoxLayout(self.alerts_tab)
        self.alert_history_table = QTableWidget(0, 3)
        self.alert_history_table.setHorizontalHeaderLabels(["Время", "Компонент", "Сообщение"])
        self.alert_history_table.horizontalHeader().setStretchLastSection(True)
        self.alert_history_table.setSortingEnabled(True)
        layout.addWidget(self.alert_history_table)

    def update_alerts_history_view(self):
        self.alert_history_table.setSortingEnabled(False)
        self.alert_history_table.setRowCount(len(self.alert_history))
        for i, alert in enumerate(self.alert_history):
            self.alert_history_table.setItem(i, 0, QTableWidgetItem(alert['Время']))
            self.alert_history_table.setItem(i, 1, QTableWidgetItem(alert['Компонент']))
            self.alert_history_table.setItem(i, 2, QTableWidgetItem(alert['Сообщение']))
        self.alert_history_table.setSortingEnabled(True)

    def init_tools_tab(self):
        layout = QVBoxLayout(self.tools_tab)
        disk_group = QGroupBox("Инструмент очистки диска")
        disk_layout = QVBoxLayout(disk_group)
        self.temp_files_check = QCheckBox("Удалить временные файлы");
        self.temp_files_check.setChecked(True)
        self.cache_files_check = QCheckBox("Очистить системный кэш(текущий пользователь)")
        disk_layout.addWidget(self.temp_files_check);
        disk_layout.addWidget(self.cache_files_check)
        cleanup_btn = QPushButton("Запустить очистку");
        cleanup_btn.clicked.connect(self.run_disk_cleanup)
        disk_layout.addWidget(cleanup_btn)
        layout.addWidget(disk_group)
        diag_group = QGroupBox("Системная и сетевая диагностика ")
        diag_layout = QVBoxLayout(diag_group)
        btn_layout = QHBoxLayout()
        disk_check_btn = QPushButton("Проверить состояние диска");
        disk_check_btn.clicked.connect(self.check_disk_health_tool)
        ping_btn = QPushButton("Пинг тест");
        ping_btn.clicked.connect(self.run_ping_test)
        speed_test_btn = QPushButton("Проверка скорости интернета");
        speed_test_btn.clicked.connect(self.run_speed_test)
        btn_layout.addWidget(disk_check_btn);
        btn_layout.addWidget(ping_btn);
        btn_layout.addWidget(speed_test_btn)
        diag_layout.addLayout(btn_layout)
        self.tools_output = QTextEdit();
        self.tools_output.setReadOnly(True);
        self.tools_output.setFont(QFont("Courier", 9))
        diag_layout.addWidget(self.tools_output)
        layout.addWidget(diag_group)

    def check_disk_health_tool(self):
        self.tools_output.setPlainText("Running disk health check...")
        QApplication.processEvents()
        try:
            encoding = 'cp866' if platform.system() == 'Windows' else 'utf-8'
            if platform.system() == 'Windows':
                cmd = ['chkdsk', 'C:']
                result = subprocess.run(cmd, capture_output=True, text=True, shell=True, timeout=120, encoding=encoding,
                                        errors='replace')
            else:
                disk = next((p.device.rstrip('0123456789') for p in psutil.disk_partitions() if not 'loop' in p.device),
                            None)
                if not disk: raise FileNotFoundError("No suitable disk found.")
                cmd = ['smartctl', '-H', disk]
                result = subprocess.run(cmd, capture_output=True, text=True, timeout=30, encoding=encoding,
                                        errors='replace')
            output = result.stdout + "\n" + result.stderr
            self.tools_output.setPlainText(output or "Command executed but produced no output.")
        except FileNotFoundError:
            self.tools_output.setPlainText("Command failed: 'chkdsk'/'smartctl' not in PATH or not installed.")
        except subprocess.TimeoutExpired:
            self.tools_output.setPlainText("Command timed out.")
        except Exception as e:
            self.tools_output.setPlainText(f"An error occurred: {str(e)}")

    def run_disk_cleanup(self):
        paths = []
        if self.temp_files_check.isChecked(): paths.append(tempfile.gettempdir())
        if self.cache_files_check.isChecked():
            paths.append(os.path.join(os.environ.get('LOCALAPPDATA', ''),
                                      'Temp') if platform.system() == 'Windows' else os.path.expanduser('~/.cache'))
        if not paths: QMessageBox.information(self, "Cleanup", "No cleanup options selected."); return
        total_files, total_size, log = 0, 0, ["Starting disk cleanup..."]
        for path in filter(None, paths):
            if not os.path.exists(path): log.append(f"\nSkipping non-existent path: {path}"); continue
            log.append(f"\nCleaning directory: {path}")
            files, size = 0, 0
            for root, _, f_names in os.walk(path):
                for f_name in f_names:
                    try:
                        f_path = os.path.join(root, f_name)
                        f_size = os.path.getsize(f_path)
                        os.remove(f_path)
                        files += 1;
                        size += f_size
                    except OSError:
                        continue
            total_files += files;
            total_size += size
            log.append(f"Deleted {files} files ({(size / 1024 ** 2):.2f} MB)")
        log.append(f"\nОчистка завершена. Всего файлов: {total_files}. Всего места: {(total_size / 1024 ** 2):.2f} MB.")
        self.tools_output.setPlainText("\n".join(log))
        QMessageBox.information(self, "Очистка завершена", f"Freed {(total_size / 1024 ** 2):.2f} MB of space.")

    def run_ping_test(self):
        target = "8.8.8.8"
        self.tools_output.setPlainText(f"Pinging {target}...")
        QApplication.processEvents()
        try:
            # Определяем параметр для количества запросов и кодировку для ОС
            param = '-n' if platform.system().lower() == 'windows' else '-c'
            encoding = 'cp866' if platform.system().lower() == 'windows' else 'utf-8'

            cmd = ['ping', param, '4', target]

            # Запускаем процесс с правильной кодировкой
            result = subprocess.run(cmd, capture_output=True, text=True, encoding=encoding, errors='replace',
                                    timeout=15)

            self.tools_output.setPlainText(result.stdout + "\n" + result.stderr)
        except Exception as e:
            self.tools_output.setPlainText(f"An error occurred: {str(e)}")

    def run_speed_test(self):
        if not self.speed_test_thread.isRunning():
            self.tools_output.setPlainText("Запуск теста скорости интренета...")
            self.speed_test_thread.result_ready.connect(self.tools_output.setPlainText)
            self.speed_test_thread.start()
        else:
            QMessageBox.warning(self, "В процессе", "Тест скорости завершен.")

    def init_settings_tab(self):
        layout = QVBoxLayout(self.settings_tab)
        tabs, general_tab, general_layout = QTabWidget(), QWidget(), QFormLayout()
        self.poll_map = {"1 сек": 1000, "2 сек": 2000, "5 сек": 5000, "10 сек": 10000,
                         "30 сек": 30000}
        self.poll_interval_combo = QComboBox();
        self.poll_interval_combo.addItems(self.poll_map.keys())
        general_layout.addRow("Период:", self.poll_interval_combo)
        general_tab.setLayout(general_layout)
        tabs.addTab(general_tab, "Главные")
        alert_tab, alert_layout = QWidget(), QFormLayout()
        self.cpu_temp_spin = QSpinBox();
        self.cpu_temp_spin.setRange(50, 120);
        self.cpu_temp_spin.setSuffix(" °C")
        self.gpu_temp_spin = QSpinBox();
        self.gpu_temp_spin.setRange(50, 120);
        self.gpu_temp_spin.setSuffix(" °C")
        self.ram_threshold_spin = QSpinBox();
        self.ram_threshold_spin.setRange(50, 100);
        self.ram_threshold_spin.setSuffix(" %")
        self.disk_threshold_spin = QSpinBox();
        self.disk_threshold_spin.setRange(50, 100);
        self.disk_threshold_spin.setSuffix(" %")
        self.popup_alerts_check = QCheckBox("Показать недавние уведомления")
        alert_layout.addRow("ЦПУ Темп ограничение:", self.cpu_temp_spin)
        alert_layout.addRow("ГПУ Темп ограничение:", self.gpu_temp_spin)
        alert_layout.addRow("ОЗУ Нагруз ограничение:", self.ram_threshold_spin)
        alert_layout.addRow("Диск Нагруз ограничение:", self.disk_threshold_spin)
        alert_layout.addRow(self.popup_alerts_check)
        alert_tab.setLayout(alert_layout)
        tabs.addTab(alert_tab, "Уведомления")
        layout.addWidget(tabs)
        save_btn = QPushButton("Сохранить настройки");
        save_btn.clicked.connect(self.save_settings)
        layout.addWidget(save_btn, 0, Qt.AlignRight)
        self.load_settings_to_ui()

    def load_settings_to_ui(self):
        rev_map = {v: k for k, v in self.poll_map.items()}
        self.poll_interval_combo.setCurrentText(rev_map.get(self.settings.get('poll_interval'), "2 seconds"))
        self.cpu_temp_spin.setValue(self.settings.get('cpu_temp_threshold', 80))
        self.gpu_temp_spin.setValue(self.settings.get('gpu_temp_threshold', 85))
        self.ram_threshold_spin.setValue(self.settings.get('ram_threshold', 90))
        self.disk_threshold_spin.setValue(self.settings.get('disk_threshold', 90))
        self.popup_alerts_check.setChecked(self.settings.get('popup_alerts', True))

    def save_settings(self):
        try:
            self.settings['poll_interval'] = self.poll_map[self.poll_interval_combo.currentText()]
            self.settings['cpu_temp_threshold'] = self.cpu_temp_spin.value()
            self.settings['gpu_temp_threshold'] = self.gpu_temp_spin.value()
            self.settings['ram_threshold'] = self.ram_threshold_spin.value()
            self.settings['disk_threshold'] = self.disk_threshold_spin.value()
            self.settings['popup_alerts'] = self.popup_alerts_check.isChecked()
            self.save_settings_to_file()
            self.apply_settings()
            QMessageBox.information(self, "Настройки",
                                    "Настройки обновлены")
        except Exception as e:
            QMessageBox.warning(self, "Ошибка", f"Не удалось установить новые настройки: {e}")

    def apply_settings(self):
        pass

    def _open_file(self, filename):
        try:
            if platform.system() == "Windows":
                os.startfile(filename)
            elif platform.system() == "Darwin":
                subprocess.run(['open', filename], check=True)
            else:
                subprocess.run(['xdg-open', filename], check=True)
        except Exception as e:
            print(f"Could not open file automatically: {e}")

    def init_reports_tab(self):
        layout = QVBoxLayout(self.reports_tab)
        group = QGroupBox("Генерация отчета")
        form_layout = QFormLayout(group)
        self.report_type_combo = QComboBox()
        self.report_type_combo.addItems(["Все показатели системы"])
        form_layout.addRow("Формат отчета:", self.report_type_combo)
        group.setLayout(form_layout)
        button_layout = QHBoxLayout()
        pdf_button = QPushButton("Сгенерировать в PDF")
        pdf_button.clicked.connect(self.generate_pdf_report)
        xml_button = QPushButton("Сгенерировать в XML")
        xml_button.clicked.connect(self.generate_xml_report)
        excel_button = QPushButton("Сгенерировать в Excel")
        excel_button.clicked.connect(self.generate_excel_report)
        button_layout.addWidget(pdf_button)
        button_layout.addWidget(xml_button)
        button_layout.addWidget(excel_button)
        layout.addWidget(group)
        layout.addLayout(button_layout)
        layout.addStretch()

    def generate_pdf_report(self):
        # 1. Проверка на наличие библиотеки и данных
        try:
            from reportlab.platypus import SimpleDocTemplate
        except ImportError:
            QMessageBox.critical(self, "Ошибка зависимости",
                                 "Библиотека 'reportlab' не установлена. Создание PDF-отчёта невозможно.\n"
                                 "Пожалуйста, выполните: pip install reportlab")
            return

        if not self.historical_data:
            QMessageBox.warning(self, "Нет данных", "Собрано недостаточно данных для создания отчёта.")
            return

        try:
            filename = os.path.join(tempfile.gettempdir(),
                                    f"system_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")

            doc = SimpleDocTemplate(filename, pagesize=letter)
            styles = getSampleStyleSheet()
            elements = []

            # --- Заголовок ---
            elements.append(Paragraph("Отчёт о состоянии системы", styles['Title']))
            elements.append(
                Paragraph(f"Сгенерировано: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
            elements.append(Spacer(1, 24))

            # --- Информация о системе ---
            elements.append(Paragraph("Информация о системе", styles['h2']))
            uname = platform.uname()
            mem = psutil.virtual_memory()
            sys_info_data = [
                ['Операционная система:', f"{uname.system} {uname.release}"],
                ['Имя компьютера:', uname.node],
                ['Процессор:', uname.processor],
                ['Всего ОЗУ:', f"{(mem.total / 1e9):.2f} ГБ"],
            ]
            sys_info_table = Table(sys_info_data, colWidths=[120, None])
            sys_info_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('TOPPADDING', (0, 0), (-1, -1), 6)
            ]))
            elements.append(sys_info_table)
            elements.append(Spacer(1, 24))

            # --- Функция для создания графика в памяти ---
            def create_chart_in_memory(figure):
                buf = io.BytesIO()
                figure.savefig(buf, format='png', dpi=150)
                buf.seek(0)
                return buf

            # --- График CPU ---
            if any(v is not None for v in self.cpu_usage_points):
                elements.append(Paragraph("Производительность CPU", styles['h2']))
                cpu_chart_img = create_chart_in_memory(self.cpu_fig)
                elements.append(Image(cpu_chart_img, width=450, height=270))
                elements.append(Spacer(1, 12))

            # --- График GPU ---
            if self.monitoring_active.get('gpu', True) and any(v is not None for v in self.gpu_load_points):
                elements.append(Paragraph("Производительность GPU", styles['h2']))
                gpu_chart_img = create_chart_in_memory(self.gpu_dashboard_fig)
                elements.append(Image(gpu_chart_img, width=450, height=270))

            # --- Сборка документа ---
            doc.build(elements)

            QMessageBox.information(self, "Отчёт создан", f"PDF-отчёт сохранён в:\n{filename}")
            self._open_file(filename)

        except Exception as e:
            QMessageBox.critical(self, "Ошибка создания PDF", f"Не удалось создать PDF-отчёт: {e}")
            # Для отладки
            import traceback
            traceback.print_exc()

    def generate_xml_report(self):
        if not self.historical_data:
            QMessageBox.warning(self, "No Data", "Not enough data collected to generate a report.")
            return
        try:
            filename = os.path.join(tempfile.gettempdir(),
                                    f"system_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xml")
            root = ET.Element("SystemReport", generated_at=datetime.now().isoformat())
            info_element = ET.SubElement(root, "SystemInformation")
            uname, mem = platform.uname(), psutil.virtual_memory()
            info_data = {
                'System': uname.system, 'NodeName': uname.node,
                'Release': uname.release, 'Version': uname.version,
                'Machine': uname.machine, 'Processor': uname.processor,
                'TotalRAM_GB': f"{mem.total / 1e9:.2f}"
            }
            for key, value in info_data.items(): ET.SubElement(info_element, key).text = value
            history_element = ET.SubElement(root, "MetricsHistory")

            def add_nodes(parent, data):
                for key, value in data.items():
                    if value is None: continue
                    safe_key = ''.join(c for c in key if c.isalnum() or c in ('_', '-'))
                    if isinstance(value, dict):
                        child = ET.SubElement(parent, safe_key)
                        add_nodes(child, value)
                    else:
                        ET.SubElement(parent, safe_key).text = str(value)

            for data_point in list(self.historical_data):
                sample_element = ET.SubElement(history_element, "MetricSample",
                                               timestamp=data_point['timestamp'].isoformat())
                point_copy = data_point.copy()
                del point_copy['timestamp']
                add_nodes(sample_element, point_copy)
            rough_string = ET.tostring(root, 'utf-8')
            reparsed = minidom.parseString(rough_string)
            pretty_xml_string = reparsed.toprettyxml(indent="  ")
            with open(filename, "w", encoding='utf-8') as f:
                f.write(pretty_xml_string)
            QMessageBox.information(self, "Отчет создан", f"XML сохранен в:\n{filename}")
            self._open_file(filename)
        except Exception as e:
            QMessageBox.critical(self, "XML Report Error", f"Failed to generate XML report: {e}")
            print(f"XML generation error: {e}")

    def generate_excel_report(self):
        if not self.historical_data:
            QMessageBox.warning(self, "No Data", "Not enough data collected to generate a report.")
            return
        try:
            filename = os.path.join(tempfile.gettempdir(),
                                    f"system_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            wb = Workbook()
            ws = wb.active
            ws.title = "Metrics History"

            def flatten_dict(d, parent_key='', sep='_'):
                items = []
                for k, v in d.items():
                    new_key = parent_key + sep + k if parent_key else k
                    if isinstance(v, dict):
                        items.extend(flatten_dict(v, new_key, sep=sep).items())
                    else:
                        items.append((new_key, v))
                return dict(items)

            flat_data = []
            all_headers = ['timestamp']
            header_set = {'timestamp'}
            for data_point in self.historical_data:
                flat_point = flatten_dict(data_point)
                flat_data.append(flat_point)
                for key in flat_point.keys():
                    if key not in header_set:
                        all_headers.append(key)
                        header_set.add(key)
            ws.append(all_headers)
            for cell in ws[1]:
                cell.font = Font(bold=True)
            for flat_point in flat_data:
                row = [flat_point.get(header, "") for header in all_headers]
                ws.append(row)
            for i, column_cells in enumerate(ws.columns):
                length = max(len(str(cell.value)) for cell in column_cells)
                ws.column_dimensions[get_column_letter(i + 1)].width = length + 2
            ws.freeze_panes = 'A2'
            wb.save(filename)
            QMessageBox.information(self, "отчет сгенерирован", f"Excel сохранен в:\n{filename}")
            self._open_file(filename)
        except Exception as e:
            QMessageBox.critical(self, "Excel Report Error", f"Failed to generate Excel report: {e}")
            print(f"Excel generation error: {e}")

    def save_settings_to_file(self):
        path = os.path.join(os.path.expanduser('~'), '.system_monitor_settings.json')
        try:
            with open(path, 'w') as f:
                json.dump(self.settings, f, indent=4)
        except Exception as e:
            print(f"Error saving settings: {e}")

    def load_settings(self):
        path = os.path.join(os.path.expanduser('~'), '.system_monitor_settings.json')
        if not os.path.exists(path): return
        try:
            with open(path, 'r') as f:
                self.settings.update(json.load(f))
        except Exception as e:
            print(f"Could not load settings: {e}")

    def closeEvent(self, event):
        reply = QMessageBox.question(self, 'Exit', 'Are you sure you want to exit?', QMessageBox.Yes | QMessageBox.No,
                                     QMessageBox.No)
        if reply == QMessageBox.Yes:
            self.aux_timer.stop()
            if self.speed_test_thread.isRunning(): self.speed_test_thread.quit(); self.speed_test_thread.wait()
            self.data_collector.stop();
            self.data_collector.wait()
            self.save_settings_to_file()
            event.accept()
        else:
            event.ignore()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    window = SystemMonitorApp()
    window.show()
    sys.exit(app.exec_())