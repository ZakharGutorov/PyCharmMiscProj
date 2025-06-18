import os
import tempfile
from datetime import datetime
import xml.etree.ElementTree as ET
from xml.dom import minidom
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet
import matplotlib.pyplot as plt


def generate_pdf_report(cpu_fig, gpu_fig):
    try:
        filename = os.path.join(tempfile.gettempdir(), f"system_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")
        doc = SimpleDocTemplate(filename, pagesize=letter)
        styles = getSampleStyleSheet()
        elements = []

        # Title
        elements.append(Paragraph("System Monitoring Report", styles['Title']))
        elements.append(Spacer(1, 12))

        # CPU Chart
        cpu_path = os.path.join(tempfile.gettempdir(), "cpu_chart.png")
        cpu_fig.savefig(cpu_path)
        elements.append(Paragraph("CPU Usage and Temperature", styles['Heading2']))
        elements.append(Image(cpu_path, width=400, height=300))
        elements.append(Spacer(1, 12))

        # GPU Chart
        gpu_path = os.path.join(tempfile.gettempdir(), "gpu_chart.png")
        gpu_fig.savefig(gpu_path)
        elements.append(Paragraph("GPU Usage and Temperature", styles['Heading2']))
        elements.append(Image(gpu_path, width=400, height=300))

        doc.build(elements)
        return filename
    except Exception as e:
        raise RuntimeError(f"PDF generation failed: {str(e)}")


def generate_xml_report(historical_data):
    try:
        filename = os.path.join(tempfile.gettempdir(), f"system_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xml")
        root = ET.Element("SystemReport")

        # System Info
        sys_info = ET.SubElement(root, "SystemInfo")
        uname = platform.uname()
        ET.SubElement(sys_info, "System").text = uname.system
        ET.SubElement(sys_info, "Node").text = uname.node
        ET.SubElement(sys_info, "Release").text = uname.release
        ET.SubElement(sys_info, "Version").text = uname.version
        ET.SubElement(sys_info, "Machine").text = uname.machine
        ET.SubElement(sys_info, "Processor").text = uname.processor

        # Metrics
        metrics = ET.SubElement(root, "Metrics")
        for data in historical_data:
            sample = ET.SubElement(metrics, "Sample", timestamp=data['timestamp'].isoformat())
            if 'cpu' in data:
                cpu = ET.SubElement(sample, "CPU")
                ET.SubElement(cpu, "Usage").text = str(data['cpu'].get('percent', ''))
                ET.SubElement(cpu, "Temperature").text = str(data['cpu'].get('temperature', ''))

            if 'memory' in data and 'virtual' in data['memory']:
                mem = ET.SubElement(sample, "Memory")
                ET.SubElement(mem, "Used").text = str(data['memory']['virtual'].get('used', ''))
                ET.SubElement(mem, "Total").text = str(data['memory']['virtual'].get('total', ''))
                ET.SubElement(mem, "Percent").text = str(data['memory']['virtual'].get('percent', ''))

        # Format and save
        xml_str = ET.tostring(root, 'utf-8')
        reparsed = minidom.parseString(xml_str)
        pretty_xml = reparsed.toprettyxml(indent="  ")

        with open(filename, 'w', encoding='utf-8') as f:
            f.write(pretty_xml)

        return filename
    except Exception as e:
        raise RuntimeError(f"XML generation failed: {str(e)}")


def generate_excel_report(historical_data):
    try:
        filename = os.path.join(tempfile.gettempdir(), f"system_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "System Metrics"

        # Headers
        headers = ["Timestamp", "CPU Usage (%)", "CPU Temp (°C)", "Memory Used (GB)", "Memory Total (GB)",
                   "Memory Usage (%)"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header).font = Font(bold=True)

        # Data
        for row, data in enumerate(historical_data, 2):
            timestamp = data['timestamp']
            cpu = data.get('cpu', {})
            mem = data.get('memory', {}).get('virtual', {})

            ws.cell(row=row, column=1, value=timestamp.isoformat())
            ws.cell(row=row, column=2, value=cpu.get('percent'))
            ws.cell(row=row, column=3, value=cpu.get('temperature'))
            ws.cell(row=row, column=4, value=mem.get('used', 0) / (1024 ** 3) if mem else 0)
            ws.cell(row=row, column=5, value=mem.get('total', 0) / (1024 ** 3) if mem else 0)
            ws.cell(row=row, column=6, value=mem.get('percent', 0) if mem else 0)

        # Auto-size columns
        for col in range(1, len(headers) + 1):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 20

        wb.save(filename)
        return filename
    except Exception as e:
        raise RuntimeError(f"Excel generation failed: {str(e)}")